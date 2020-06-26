#output.xlsx
#
#where output.xlsx is the unified file

#This works FROM/TO the xlsx format. Libreoffice might help to convert from xls.
#localc --headless  --convert-to xlsx somefile.xls

import sys
from copy import copy

from openpyxl import load_workbook,Workbook

def createNewWorkbook(manyWb):
    for wb in manyWb:
        for sheetName in wb.sheetnames:
            o = theOne.create_sheet(sheetName)
            safeTitle = o.title
            copySheet(wb[sheetName],theOne[safeTitle])

def copySheet(sourceSheet,newSheet):
    for row in sourceSheet.rows:
        for cell in row:
            newCell = newSheet.cell(row=cell.row, column=cell.col_idx,
                    value= cell.value)
            if cell.has_style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = copy(cell.number_format)
                newCell.protection = copy(cell.protection)
                newCell.alignment = copy(cell.alignment)

#filesInput = sys.argv[1:]
#theOneFile = filesInput.pop(-1)
#myfriends = [ load_workbook(f) for f in filesInput ]

#try this if you are bored
#myfriends = [ openpyxl.load_workbook(f) for k in range(200) for f in filesInput ]

theOne = Workbook()
del theOne['Sheet'] #We want our new book to be empty. Thanks.
createNewWorkbook(myfriends)
theOne.save(theOneFile)
